import re
import pandas as pd
import os
import string
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from modules.bnk import extractBNKFile

def process_string(s='', Rvalues = []):
    # Add new feature: replace the four digits after "PL" with "xxx0"
    s = re.sub(r'(PL\d{3})\d', r'\g<1>0', s)

    pattern1 = r'([a-zA-Z])(_((\d+_)*\d+))$'

    match1 = re.search(pattern1, s)

    if match1:
        index = match1.start(2)
        second_part = s[index:]
        numbers = second_part.split('_')[1:]  # split the string and remove the first empty string
        if len(numbers) >= 2:
            second_part = '_' + numbers[0]
        else:
            second_part = ''
        s = (s[:index] + second_part)

    # Backup the string s
    original_s = s
    # Compare with the string s
    while '_' in s:
        s = s[:s.rfind('_')]
        if s in Rvalues:
            return s

    return original_s

def Event2wemID_Process():
    excel_path = r'.\\Yours\\GBFR#2Motion_EventName_ComparisonTable.xlsx'
    # Read the excel file
    df = pd.read_excel(excel_path)

    # Extract all column values in the form of "PL_vo_Event_number"
    columns = [col for col in df.columns if 'PL_vo_Event_' in col]
    values = df[columns].values.flatten()

    # Remove invalid and duplicate values
    values = [value for value in values if pd.notna(value)]
    values = list(set(values))

    # 获取所有以 "vo_pl" 开头的 bnk 或 pck 文件
    folder_path = ".\\Yours\\"
    # files = [f for f in os.listdir(folder_path) if f.startswith("vo_pl") and (f.endswith(".bnk") or f.endswith(".pck"))]
    files = [f for f in os.listdir(folder_path) if f.startswith("vo_pl") and (f.endswith(".bnk"))]
    # 对每个文件执行命令行
    for file in files:
        # 构造解包完 wem 文件的存储路径
        output_folder = os.path.join(".\\Yours\\wem_org\\", file)

        # 检查存储路径是否存在，如果不存在就创建它
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # 根据文件类型选择正确的工具
        if file.endswith(".bnk"):
            # tool = ".\\Yours\\nier_BNK_Util\\nier_BNK_Util.py"
            # cmd = f"{tool} \"{os.path.join(folder_path, file)}\" -e \"{output_folder}\""  # 构造命令行
            # subprocess.run(cmd, shell=True)  # 执行命令行# 执行命令行
            SOURCEFILE = f"{os.path.join(folder_path, file)}"
            OUTPUTDIR = f"{output_folder}"
            extractBNKFile(SOURCEFILE, OUTPUTDIR)

        else:  # file.endswith(".pck")
            # tool = ".\\Yours\\replicant_PCK_Util\\replicant_PCK_Util.py"
            print(fr"Please use RingingBloom to unpack .pck to .\Yours\wem_org\(.pck file full name) folder")

    # # 获取所有以 ".pck" 结尾的文件夹
    # folder_path = ".\\Yours\\wem_org\\"
    # folders = [f for f in os.listdir(folder_path) if f.endswith(".pck") and os.path.isdir(os.path.join(folder_path, f))]
    # # 对每个文件夹执行操作
    # for folder in folders:
    #     # 构造 "japanese" 文件夹的路径
    #     japanese_folder = os.path.join(folder_path, folder, "japanese")
    #
    #     # 检查 "japanese" 文件夹是否存在
    #     if os.path.exists(japanese_folder):
    #         # 移动 "japanese" 文件夹中的所有文件到它的父文件夹
    #         for filename in os.listdir(japanese_folder):
    #             src_path = os.path.join(japanese_folder, filename)
    #             dst_path = os.path.join(folder_path, folder, filename)
    #
    #             # 移动文件，如果目标路径存在同名文件，直接覆盖
    #             shutil.move(src_path, dst_path)
    #
    #         # 删除 "japanese" 文件夹
    #         os.rmdir(japanese_folder)
    #
    #     # 构造 "sfx" 文件夹的路径
    #     sfx_folder = os.path.join(folder_path, folder, "sfx")
    #
    #     # 检查 "sfx" 文件夹是否存在
    #     if os.path.exists(sfx_folder):
    #         # 删除 "sfx" 文件夹
    #         os.rmdir(sfx_folder)

    # 使用正则表达式匹配文件夹名
    pattern = re.compile(r'^vo_pl\d{4}.*\.(bnk|pck)$')

    # 获取所有匹配的子文件夹
    folder_path_wemorg = ".\\Yours\\wem_org\\"
    subfolders = [f for f in os.listdir(folder_path_wemorg) if os.path.isdir(os.path.join(folder_path_wemorg, f)) and pattern.match(f)]
    # 定义排序规则
    def sort_key(x):
        base_name, ext = os.path.splitext(x)
        return (ext != '.bnk', base_name.count('_'), base_name)

    # 按照规则排序子文件夹
    subfolders.sort(key=sort_key)

    print(f"{subfolders}")

    result_dict = {}
    df_list = []  # 空的 DataFrame

    result_list = []  # 使用列表来存储所有的字典
    prefixes = iter(string.ascii_lowercase)  # 创建一个迭代器，生成字母前缀
    prefix = next(prefixes)  # 初始化前缀

    for subfolder in subfolders:
        subfolder_path = os.path.join(folder_path_wemorg, subfolder)
        num_serial = 1  # 在每个子文件夹开始时重置 NumSerial
        # 获取子文件夹中的所有 .wem 文件
        filenames = [filename for filename in os.listdir(subfolder_path) if filename.endswith(".wem")]
        # 将文件名转换为整数并排序
        filenames = sorted(filenames, key=lambda x: int(os.path.splitext(x)[0]))
        for filename in filenames:
            if filename.endswith(".wem"):
                file_path = os.path.join(subfolder_path, filename)
                wemID = os.path.basename(file_path).split('.')[0]
                with open(file_path, "rb") as file:
                    hex_content = file.read().hex()
                ascii_string = bytes.fromhex(hex_content).decode("ascii", errors="ignore")
                pattern = r"PL.*?_vo.*?data"
                matches = re.findall(pattern, ascii_string)
                if not matches:
                    result_list.append({"wemID": wemID,
                                        "PL_vo_Event_WithSerial": "",
                                        "PL_vo_Event": "",
                                        "bnk_Belonging": subfolder,
                                        "NumSerial": prefix + "_" + str(num_serial)})  # 添加 NumSerial
                else:
                    clean_matches = []
                    processed_matches = []
                    for match in matches:
                        clean_match = re.sub(r"[^a-zA-Z0-9_]", "_", match)
                        clean_match = clean_match.replace("data", "")
                        clean_match = clean_match.rstrip("_")
                        clean_matches.append(clean_match)
                        processed_matches.append(process_string(s=clean_match, Rvalues = values))
                    result_list.append({"wemID": wemID,
                                        "PL_vo_Event_WithSerial": "\n".join(clean_matches),
                                        "PL_vo_Event": "\n".join(processed_matches),
                                        "bnk_Belonging": subfolder,
                                        "NumSerial": prefix + "_" + str(num_serial)})  # 添加 NumSerial
                num_serial += 1  # 更新 NumSerial
        prefix = next(prefixes)  # 更新前缀

    # 使用列表来创建 DataFrame
    df = pd.DataFrame(result_list)

    # 按照 'NumSerial' 列的值对 DataFrame 进行排序
    df['Prefix'] = df['NumSerial'].str.extract(r'(\D+)')  # 提取前缀
    df['Number'] = df['NumSerial'].str.extract(r'(\d+)').astype(int)  # 提取数字
    # df.sort_values(by=['Prefix', 'Number'], ascending=[False, True], inplace=True)
    df.sort_values(by=['Prefix', 'Number'], ascending=[True, True], inplace=True)
    df.drop(columns=['Prefix', 'Number'], inplace=True)  # 删除临时列

    # 将 'NumSerial' 列移动到第一列
    df = df.reindex(columns=['NumSerial'] + [col for col in df.columns if col != 'NumSerial'])

    # 将结果保存到 Excel 文件
    output_file = os.path.join(r'.\Yours', "GBFR#3EventName_wemID_ComparisonTable.xlsx")
    df.to_excel(output_file, sheet_name="EventName2wemID", index=False, engine='openpyxl')


    # 使用 openpyxl 打开 Excel 文件并设置单元格的对齐方式和文本换行
    wb = load_workbook(r'.\Yours\GBFR#3EventName_wemID_ComparisonTable.xlsx')
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            # cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb.save(r'.\Yours\GBFR#3EventName_wemID_ComparisonTable.xlsx')
    print(f'Excel updated data saved in {output_file}')

if __name__ == "__main__":
    Event2wemID_Process()