import os
import pandas as pd
import re
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import json
from openpyxl import load_workbook
from collections import defaultdict
import tkinter as tk
from tkinter import messagebox

def convert_value(x):
    if pd.isnull(x):  # 如果数据是空的
        return ''
    elif isinstance(x, float) and x.is_integer():  # 如果数据是整数的浮点数
        x = int(x)  # 转换为整数
        if x < 1000:  # 如果数字小于1000
            return str(x).zfill(4)  # 转换为字符串，并确保它是4位的
        else:
            return str(x)  # 否则，直接转换为字符串
    else:
        return x  # 其他情况，保持原样

def fdColIdx(worksheet, col_name):
    col_idx = None
    for i, cell in enumerate(worksheet[1], start=1):
        if cell.value == col_name:
            col_idx = i
            break
    return  col_idx

def ConfigSheet(worksheet):
    # 设置单元格的格式
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 遍历所有列，设置列宽
    for i, column in enumerate(worksheet.columns, start=1):
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if cell.value[0] != "=":
                    if "\n" in cell.value:
                        lengths = [len(chunk) for chunk in cell.value.split("\n")]
                        max_length = max(max_length, max(lengths))
                    else:
                        max_length = max(max_length, len(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[get_column_letter(i)].width = adjusted_width

    # 把第一行冻结，并加粗字体
    worksheet.freeze_panes = 'A2'
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

def MixUp_Process():
    # region 第1步：找到与这个脚本在同一个路径下的Excel文件
    file_names = ["GBFR#1Action_Motion_ComparisonTable.xlsx", "GBFR#2Motion_EventName_ComparisonTable.xlsx", "GBFR#3EventName_wemID_ComparisonTable.xlsx"]
    dfs = [pd.read_excel(os.path.join(r'.\Yours', f'{file_name}')) for file_name in file_names]
    # endregion

    # 第2步：新建一个Excel文件
    with pd.ExcelWriter(r'.\Yours\GBFR#0AllMixUp_PlayerVoice.xlsx', engine='openpyxl') as writer:
        for col in dfs[0].columns:
            dfs[0][col] = dfs[0][col].apply(convert_value)
        dfs[0].to_excel(writer, sheet_name='Act2Mot', index=False)
        for col in dfs[1].columns:
            dfs[1][col] = dfs[1][col].apply(convert_value)
        dfs[1].to_excel(writer, sheet_name='Mot2Event', index=False)
        dfs[2].to_excel(writer, sheet_name='Event2wemID', index=False)

        # 第3步：新建一个sheet叫“MixUp”
        df_mixup = pd.DataFrame(columns=['ActioinName_DIY', 'ActionID', 'MotionID', 'PL_vo_Event', 'PL_vo_Event_WithSerial', 'wemID', 'NumSerial', 'bnk_Belonging'])
        df_mixup.to_excel(writer, sheet_name='MixUp', index=False)

    # region 参数
    wb = load_workbook(r'.\Yours\GBFR#0AllMixUp_PlayerVoice.xlsx')
    ws_act2mot = wb['Act2Mot']
    ws_mot2event = wb['Mot2Event']
    ws_event2wemid = wb['Event2wemID']
    ws_mixup = wb['MixUp']
    separator1 = "- - - - - - - - - - - -"# 定义特殊分隔符
    separator2 = ".. .. .. .. .."
    Nansignal = "#"
    col_idx_pveWs_E2i = fdColIdx(ws_event2wemid,'PL_vo_Event_WithSerial')
    col_idx_wemID_E2i = fdColIdx(ws_event2wemid,'wemID')
    col_idx_NumSerial_E2i = fdColIdx(ws_event2wemid,'NumSerial')
    col_idx_bnkBelonging_E2i = fdColIdx(ws_event2wemid,'bnk_Belonging')
    col_idx_TrNDIY_mx = fdColIdx(ws_mixup,'ActioinName_DIY')
    col_idx_aID_mx = fdColIdx(ws_mixup,'ActionID')
    col_idx_mID_mx = fdColIdx(ws_mixup,'MotionID')
    col_idx_pve_mx = fdColIdx(ws_mixup,'PL_vo_Event')
    col_idx_pveWs_mx = fdColIdx(ws_mixup,'PL_vo_Event_WithSerial')
    col_idx_wemID_mx = fdColIdx(ws_mixup,'wemID')
    col_idx_NumSerial_mx = fdColIdx(ws_mixup,'NumSerial')
    col_idx_bnkBelonging_mx = fdColIdx(ws_mixup,'bnk_Belonging')
    # endregion

    # region 第4步：遍历sheet“Act2Mot”里的ActionID列下的所有单元格
    header = [cell.value for cell in ws_act2mot[1]]# 获取标题行
    action_motion_dict = {}# 创建一个字典来存储每个ActionID对应的MotionID列表
    for i, row in enumerate(ws_act2mot.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):
        action_id = row[0]
        motion_ids = []
        for j, cell in enumerate(ws_act2mot[i], start=1):
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            # 检查标题是否匹配
            if re.match(r'MotionID_\d+', str(header[j-1])):
                motion_ids.append(str(cell.value))
        # 将MotionID列表添加到字典中
        action_motion_dict[action_id] = motion_ids
    # endregion

    # region 第5步：遍历sheet“Mot2Event”里的MotionID列下的所有单元格
    header = [cell.value for cell in ws_mot2event[1]]# 获取标题行
    motion_event_dict = {}# 创建一个字典来存储每个MotionID对应的PL_vo_Event列表
    for i, row in enumerate(ws_mot2event.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):
        motion_id = row[0]
        event_ids = []
        for j, cell in enumerate(ws_mot2event[i], start=1):
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            # 检查标题是否匹配
            if re.match(r'PL_vo_Event_\d+', str(header[j-1])):
                event_ids.append(str(cell.value))
        # 将PL_vo_Event列表添加到字典中
        motion_event_dict[motion_id] = event_ids
    # endregion

    # region 第6步：遍历sheet“Event2wemID”里的PL_vo_Event列下的非空单元格
    header = [cell.value for cell in ws_event2wemid[1]]# 获取标题行
    # 获取PL_vo_Event列的索引
    col_idx_event = None
    for i, cell in enumerate(ws_event2wemid[1], start=1):
        if cell.value == 'PL_vo_Event':
            col_idx_event = i
            break

    # 获取PL_vo_Event_WithSerial列的索引
    col_idx_serial = None
    for i, cell in enumerate(ws_event2wemid[1], start=1):
        if cell.value == 'PL_vo_Event_WithSerial':
            col_idx_serial = i
            break

    # 创建一个字典来存储每个PL_vo_Event对应的PL_vo_Event_WithSerial列表
    event_serial_dict = {}
    for i, row in enumerate(ws_event2wemid.iter_rows(min_row=2, min_col=col_idx_event, max_col=col_idx_event, values_only=True), start=2):
        event_id = row[0]
        if event_id is not None and event_id not in event_serial_dict:
            event_serials = []
            for j, row in enumerate(ws_event2wemid.iter_rows(min_row=2, values_only=True), start=2):
                if ws_event2wemid.cell(row=j, column=col_idx_event).value == event_id:
                    cell_value = ws_event2wemid.cell(row=j, column=col_idx_serial).value
                    if cell_value is not None:
                        event_serials.append(cell_value)
            # 将PL_vo_Event_WithSerial列表添加到字典中
            event_serial_dict[event_id] = event_serials
    # endregion

    # region 6.1：提取CharacterID
    character_ids = [] # 创建一个空列表来保存CharacterID
    # 遍历bnk_Belonging列
    for column in ws_event2wemid.iter_cols(1, ws_event2wemid.max_column):
        if column[0].value == "bnk_Belonging":
            column_letter = column[0].column_letter
            break

    # 现在你可以使用这个字母标识来遍历列
    for cell in ws_event2wemid[column_letter]:
        # 使用正则表达式提取“vo_pl”后面的四个数字
        match = re.search(r'vo_pl(\d{4})', cell.value)
        if match:
            # 如果找到匹配项，将其添加到列表中
            character_ids.append(match.group(1))

    character_ids = list(set(character_ids))# 去除重复的CharacterID

    # 如果列表中只有一个CharacterID，打印它
    if len(character_ids) == 1:
        CharID = f"Pl{character_ids[0]}"
    # 如果列表中有多个CharacterID，抛出异常
    elif len(character_ids) > 1:
        raise Exception("Multiple CharacterIDs found. The CharacterIDs are inconsistent.")
    # endregion

    # region 6.2:提取GBFR logs里的Json文件
    CharName = ''
    try:
        with open('./Yours/ui.json', 'r', encoding='utf-8') as f:
            uiData = json.load(f)
        if 'characters' in uiData and CharID in uiData['characters']:
            CharName = uiData['characters'][CharID]
        else:
            print(f'No character found with ID: {CharID}')
    except FileNotFoundError:
        print('Warning: ui.json not found. Continuing execution.')

    # region 7: 根据action_motion_dict写入mixup的ActionID和MotionID
    # 遍历字典并写入数据
    for action_id, motion_ids in action_motion_dict.items():

        motion_ids_str = '\n'.join(motion_ids)# 将motion_ids列表转换为字符串，元素之间用'\n'分隔
        mixUpMaxRow = ws_mixup.max_row+1
        ws_mixup.cell(mixUpMaxRow, col_idx_aID_mx).value = action_id
        ws_mixup.cell(mixUpMaxRow, col_idx_mID_mx).value = motion_ids_str

    wb.save(r'.\Yours\GBFR#0AllMixUp_PlayerVoice.xlsx')
    # endregion

    # region 8：根据motion_event_dict写入mixup的PL_vo_Event
    # 遍历字典并写入数据
    for i, row in enumerate(ws_mixup.iter_rows(min_row=2, min_col=col_idx_mID_mx, max_col=col_idx_mID_mx, values_only=True), start=2):
        motion_id = row[0]
        # 检查motion_id是否包含多行
        if motion_id is not None and '\n' in motion_id:
            # 如果包含多行，那么将它们分割成列表
            motion_ids = motion_id.split('\n')
            # 遍历列表并写入数据
            event_ids_all = []
            for motion_id in motion_ids:
                if motion_id in motion_event_dict:
                    if not motion_event_dict.get(motion_id, []):
                        event_ids_all.append(Nansignal)
                    else:
                        # 将event_ids列表转换为字符串，元素之间用'\n'分隔
                        event_ids_str = '\n'.join(motion_event_dict[motion_id])
                        event_ids_all.append(event_ids_str)
                else:
                    event_ids_all.append(Nansignal)
            # 将所有的event_ids_str用特殊分隔符连接起来，并写入数据
            ws_mixup.cell(row=i, column=col_idx_pve_mx).value = f"\n{separator1}\n".join(event_ids_all)
        else:
            if motion_id in motion_event_dict:
                # 将event_ids列表转换为字符串，元素之间用'\n'分隔
                event_ids_str = f"\n{separator1}\n".join(motion_event_dict[motion_id])
                # 写入数据
                ws_mixup.cell(row=i, column=col_idx_pve_mx).value = event_ids_str
    # endregion

    # region 9：根据event_serial_dict写入mixup的PL_vo_Event_WithSerial
    # 遍历字典并写入数据
    for i, row in enumerate(ws_mixup.iter_rows(min_row=2, min_col=col_idx_pve_mx, max_col=col_idx_pve_mx, values_only=True), start=2):
        event_id = row[0]
        # 检查event_id是否包含多行
        if event_id is not None and '\n' in event_id:
            # 如果包含多行，那么将它们分割成列表
            event_ids = event_id.split('\n')
            # 遍历列表并写入数据
            event_serials_all = []
            for index, event_id in enumerate(event_ids):
                if event_id != Nansignal and event_id != separator1:
                    if event_id in event_serial_dict:
                        if not event_serial_dict.get(event_id, []):
                            event_serials_all.append(Nansignal)
                        else:
                            # 将event_serials列表转换为字符串，元素之间用'\n'分隔
                            event_serials_str = '\n'.join(event_serial_dict[event_id])
                            event_serials_all.append(event_serials_str)
                    else:
                        event_serials_all.append(Nansignal)
                    if index != len(event_ids) - 1:
                        event_serials_all.append(separator2)
                else:
                    if event_id == separator1 and event_serials_all[len(event_serials_all)-1] == separator2:
                        event_serials_all.pop(len(event_serials_all)-1)
                    event_serials_all.append(event_id)
            # 将所有的event_serials_str用'\n'连接起来，并写入数据
            ws_mixup.cell(row=i, column=col_idx_pveWs_mx).value = '\n'.join(event_serials_all)
        else:
            if event_id in event_serial_dict:
                # 将event_serials列表转换为字符串，元素之间用'\n'分隔
                event_serials_str = '\n'.join(event_serial_dict[event_id])
                # 写入数据
                ws_mixup.cell(row=i, column=col_idx_pveWs_mx).value = event_serials_str

    # 设置单元格的格式
    for row in ws_mixup.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    #10
    # 遍历“PL_vo_Event_WithSerial”列并写入数据
    for i, row in enumerate(ws_mixup.iter_rows(min_row=2, min_col=col_idx_pveWs_mx, max_col=col_idx_pveWs_mx, values_only=True), start=2):
        event_id = row[0]
        # 检查event_id是否包含多行
        if event_id is not None and '\n' in event_id:
            # 如果包含多行，那么将它们分割成列表
            event_ids = event_id.split('\n')
            # 遍历列表并写入数据
            wemid_all = []
            wemid_all_n = []
            wemid_all_bb = []
            for index, event_id in enumerate(event_ids):
                if event_id != Nansignal and event_id != separator1 and event_id != separator2:
                    # 在“Event2wemID”表中查找“PL_vo_Event_WithSerial”列的值
                    for j, row_event2wemid in enumerate(ws_event2wemid.iter_rows(min_row=2, min_col=col_idx_wemID_E2i, max_col=col_idx_wemID_E2i, values_only=True), start=1):
                        if event_id == ws_event2wemid.cell(row=j, column=col_idx_pveWs_E2i).value:
                            wemid_all.append(ws_event2wemid.cell(row=j, column=col_idx_wemID_E2i).value)
                            wemid_all_n.append(ws_event2wemid.cell(row=j, column=col_idx_NumSerial_E2i).value)
                            wemid_all_bb.append(ws_event2wemid.cell(row=j, column=col_idx_bnkBelonging_E2i).value)
                else:
                    wemid_all.append(event_id)
                    wemid_all_n.append(event_id)
                    wemid_all_bb.append(event_id)
            # 将所有的wemid用'\n'连接起来，并写入数据
            ws_mixup.cell(row=i, column=col_idx_wemID_mx).value = '\n'.join(map(str, wemid_all))
            ws_mixup.cell(row=i, column=col_idx_NumSerial_mx).value = '\n'.join(map(str, wemid_all_n))
            ws_mixup.cell(row=i, column=col_idx_bnkBelonging_mx).value = '\n'.join(map(str, wemid_all_bb))
        else:
            if event_id is not None:
                # 在“Event2wemID”表中查找“PL_vo_Event_WithSerial”列的值
                for j, row_event2wemid in enumerate(ws_event2wemid.iter_rows(min_row=2, min_col=col_idx_wemID_E2i, max_col=col_idx_wemID_E2i,values_only=True), start=1):
                    if event_id == ws_event2wemid.cell(row=j, column=col_idx_pveWs_E2i).value:
                        # 写入数据
                        ws_mixup.cell(row=i, column=col_idx_wemID_mx).value = ws_event2wemid.cell(row=j, column=col_idx_wemID_E2i).value
                        ws_mixup.cell(row=i, column=col_idx_NumSerial_mx).value = ws_event2wemid.cell(row=j, column=col_idx_NumSerial_E2i).value
                        ws_mixup.cell(row=i, column=col_idx_bnkBelonging_mx).value = ws_event2wemid.cell(row=j, column=col_idx_bnkBelonging_E2i).value

    # 11
    # 读取NumSerial、wemID和PL_vo_Event_WithSerial列的所有值并合并成一个列表
    columns = [col_idx_bnkBelonging_mx, col_idx_NumSerial_mx, col_idx_wemID_mx, col_idx_pveWs_mx]

    values_mx = {col: [] for col in columns}
    values_prepare = {col: [] for col in columns}

    for row in ws_mixup.iter_rows(min_row=2, max_col=max(columns), values_only=True):
        for col in columns:
            cell_value = row[col - 1]  # 列索引从0开始，所以需要减1
            if cell_value is not None and isinstance(cell_value, str):
                values_mx[col].extend(cell_value.split('\n'))

    for col in columns: # 去除特殊分隔符和Nansignal
        values_mx[col] = [value for value in values_mx[col] if value not in [separator1, separator2, Nansignal]]

    for col in columns: # 去除重复的值
        values_mx[col] = list(set(values_mx[col]))

    # 获取values中的wemID列的值
    # values_wemID = set(values_mx.get(col_idx_wemID_mx, []))
    values_NumS = set(values_mx.get(col_idx_wemID_mx, []))
    # 遍历ws_event2wemid工作表的每一行
    for row in ws_event2wemid.iter_rows(min_row=2, max_col=max([col_idx_bnkBelonging_E2i, col_idx_pveWs_E2i, col_idx_wemID_E2i, col_idx_NumSerial_E2i]), values_only=True):
        NumS_value = row[col_idx_NumSerial_E2i - 1]  # 列索引从0开始，所以需要减1
        if NumS_value is not None and str(NumS_value) not in values_NumS:
            # 如果wemID_value存在于ws_event2wemid的wemID列而不存在于values.get(col_idx_wemID_mx)里的值
            # 则将对应的值添加到values_prepare字典中
            values_prepare[col_idx_pveWs_mx].append(row[col_idx_pveWs_E2i - 1])
            values_prepare[col_idx_wemID_mx].append(row[col_idx_wemID_E2i - 1])
            values_prepare[col_idx_NumSerial_mx].append(row[col_idx_NumSerial_E2i - 1])
            values_prepare[col_idx_bnkBelonging_mx].append(row[col_idx_bnkBelonging_E2i - 1])

    valid_rows = ws_mixup.max_row# 找出ws_mixup的有效行数
    # 将处理后的列表填写到NumSerial、wemID和PL_vo_Event_WithSerial列下的有效行下
    for col in columns:
        for i, values in enumerate(values_prepare[col], start=valid_rows+1):
            ws_mixup.cell(row=i, column=col, value=values)

    old_max_row = valid_rows# 记录添加新行之前的行数
    valid_rows = ws_mixup.max_row# 更新ws_mixup的有效行数

    # 遍历PL_vo_Event_WithSerial列的每个单元格
    for i, row in enumerate(ws_mixup.iter_rows(min_row=old_max_row+1, max_row=valid_rows, max_col=col_idx_pveWs_mx, values_only=True), start=old_max_row+1):
        cell_value = row[col_idx_pveWs_mx - 1]  # 列索引从0开始，所以需要减1
        if cell_value is not None and isinstance(cell_value, str):
            # 在event_serial_dict字典中逆向查找对应的键值
            for key, values in event_serial_dict.items():
                if cell_value in values:
                    # 将找到的键值写入PL_vo_Event列
                    ws_mixup.cell(row=i, column=col_idx_pve_mx, value=key)
                    break  # 找到后就跳出循环

    wb.save(r'.\Yours\GBFR#0AllMixUp_PlayerVoice.xlsx')

    # 余下的排列
    df = pd.read_excel(r'.\Yours\GBFR#0AllMixUp_PlayerVoice.xlsx', sheet_name='MixUp')
    # 将DataFrame分为两部分
    df1 = df.iloc[:(old_max_row-1)]
    df2 = df.iloc[(old_max_row-1):(valid_rows-1)]
    df2 = df2.sort_values(by='PL_vo_Event_WithSerial').reset_index(drop=True)# 对df2按照'PL_vo_Event_WithSerial'列进行排序

    # 将排序后的df2写回原始Excel文件的相应位置
    start_row = old_max_row + 1  # Excel的行号从1开始，加上列标题行
    for index, row in df2.reset_index(drop=True).iterrows():
        for col_num, value in enumerate(row, start=1):
            ws_mixup.cell(row=start_row + index, column=col_num, value=value)

    wb.save(r'.\Yours\GBFR#0AllMixUp_PlayerVoice.xlsx')# 保存修改后的Excel文件

    #合并相同单元格
    previous_value = None
    start_row = None
    for row in range(old_max_row+1, valid_rows+1):
        current_value = ws_mixup[f"{get_column_letter(col_idx_pve_mx)}{row}"].value
        if current_value == previous_value:
            continue
        else:
            if start_row is not None:
                ws_mixup.merge_cells(start_row=start_row, start_column=col_idx_pve_mx, end_row=row-1, end_column=col_idx_pve_mx)
            start_row = row
            previous_value = current_value

    # 如果最后一组单元格还没有合并，进行合并
    if start_row is not None:
        ws_mixup.merge_cells(start_row=start_row, start_column=col_idx_pve_mx, end_row=valid_rows, end_column=col_idx_pve_mx)
    # endregion

    # region 10: 创建一个MixUpSep表用来查找
    ws_mixup_sep = wb.create_sheet("MixUpSep")
    headers = ["ActioinName_DIY", "ActionID", "MotionID", "PL_vo_Event", "PL_vo_Event_WithSerial", "wemID", "NumSerial", "bnk_Belonging"]
    ws_mixup_sep.append(headers)
    col_idx_TrNDIY_mxS = fdColIdx(ws_mixup_sep,'ActioinName_DIY')
    col_idx_aID_mxS = fdColIdx(ws_mixup_sep,'ActionID')
    col_idx_mID_mxS = fdColIdx(ws_mixup_sep,'MotionID')
    col_idx_pveWs_mxS = fdColIdx(ws_mixup_sep,'PL_vo_Event_WithSerial')
    col_idx_pve_mxS = fdColIdx(ws_mixup_sep,'PL_vo_Event')
    col_idx_wemID_mxS = fdColIdx(ws_mixup_sep,'wemID')
    col_idx_NumSerial_mxS = fdColIdx(ws_mixup_sep,'NumSerial')
    col_idx_bnkBelonging_mxS = fdColIdx(ws_mixup_sep,'bnk_Belonging')

    # 树状散开排列
    for action_id, motion_ids in action_motion_dict.items():
        # Iterate over the list of motion_ids
        for motion_id in motion_ids:
            # Get the corresponding events for the current motion_id
            events = motion_event_dict.get(motion_id, [])
            # Iterate over the list of events
            for event in events:
                # Get the corresponding serials for the current event
                serials = event_serial_dict.get(event, [])
                # Iterate over the list of serials
                for serial in serials:
                    # Append the data to the MixUpSep worksheet
                    # ws_mixup_sep.append([action_id, motion_id, event, serial, "", ""])
                    mixUpSepMaxRow = ws_mixup_sep.max_row+1
                    ws_mixup_sep.cell(mixUpSepMaxRow, col_idx_aID_mxS).value = action_id
                    ws_mixup_sep.cell(mixUpSepMaxRow, col_idx_mID_mxS).value = motion_id
                    ws_mixup_sep.cell(mixUpSepMaxRow, col_idx_pve_mxS).value = event
                    ws_mixup_sep.cell(mixUpSepMaxRow, col_idx_pveWs_mxS).value = serial

    old_max_row_mxs = ws_mixup_sep.max_row
    df2_values = df2.values.tolist()
    # Append the values to the MixUpSep worksheet
    for row_values in df2_values:
        ws_mixup_sep.append(row_values)

    # Create a dictionary to map PL_vo_Event_WithSerial to wemID
    event_wemid_dict_E2i = {}
    event_NumSerial_dict_E2i = {}
    event_bnkBelonging_dict_E2i = {}
    for row in ws_event2wemid.iter_rows(min_row=2, values_only=True):
        event_wemid_dict_E2i[row[col_idx_pveWs_E2i - 1]] = row[col_idx_wemID_E2i - 1]
        event_NumSerial_dict_E2i[row[col_idx_pveWs_E2i - 1]] = row[col_idx_NumSerial_E2i - 1]
        event_bnkBelonging_dict_E2i[row[col_idx_pveWs_E2i - 1]] = row[col_idx_bnkBelonging_E2i - 1]

    # Iterate over the rows in the MixUpSep worksheet
    for row in ws_mixup_sep.iter_rows(min_row=2, max_row= old_max_row_mxs, values_only=False):
        row[col_idx_wemID_mxS - 1].value = event_wemid_dict_E2i.get(row[col_idx_pveWs_mxS - 1].value)
        row[col_idx_NumSerial_mxS - 1].value = event_NumSerial_dict_E2i.get(row[col_idx_pveWs_mxS - 1].value)
        row[col_idx_bnkBelonging_mxS - 1].value = event_bnkBelonging_dict_E2i.get(row[col_idx_pveWs_mxS - 1].value)
    # endregion

    # region 12:写函数入ActionName_DIY
    for column in ws_mixup.columns:
        if column[0].value == 'ActionID':
            action_id_column = column
            break

    for row in range(2, old_max_row + 1):
        # 获取单元格的引用
        cell = ws_mixup.cell(row=row, column= (col_idx_aID_mx))
        col_letter = get_column_letter(cell.column)
        cell_reference = f'{col_letter}{cell.row}'
        table_name = "YourRecord"
        action_id = str(ws_mixup[cell_reference].value)

        if CharID in uiData['skills'] and action_id in uiData['skills'][CharID]:
            formula = uiData['skills'][CharID][action_id]
        else:
            formula = f'=IFERROR(INDEX(YourRecord!$A:$A,MATCH({cell_reference},YourRecord!$F:$F,0)),"")'

        ws_mixup.cell(row=row, column=(col_idx_TrNDIY_mx), value=formula)

    for row in range(2, ws_mixup_sep.max_row + 1):
        # 获取单元格的引用
        cell = ws_mixup_sep.cell(row=row, column=col_idx_aID_mxS)
        if cell.value is not None:
            col_letter = get_column_letter(cell.column)
            cell_reference = f'{col_letter}{cell.row}'
            formula = f'=IFERROR(INDEX(MixUp!A:A,MATCH({cell_reference},MixUp!B:B,0)),"")'
            ws_mixup_sep.cell(row=row, column=col_idx_TrNDIY_mxS, value=formula)
    # endregion

    # region 15:新建一个MixUpReverse表
    # 创建新的工作表
    ws_mixup_reverse = wb.create_sheet("MixUpReverse")

    # 复制ws_event2wemid的内容到新的工作表
    for row in ws_event2wemid.iter_rows():
        ws_mixup_reverse.append([cell.value for cell in row])

    # 在新的工作表中添加"ActionID"和"Motion"列
    ws_mixup_reverse.cell(row=1, column=ws_mixup_reverse.max_column + 1, value="ActionID")
    ws_mixup_reverse.cell(row=1, column=ws_mixup_reverse.max_column + 1, value="MotionID")

    # 创建两个字典，键是"NumSerial"，值分别是"ActionID"和"MotionID"
    numserial_to_actionid = defaultdict(list)
    numserial_to_motionid = defaultdict(list)

    # 找到"NumSerial"、"ActionID"和"MotionID"列的位置
    numserial_col = next(col_idx for col_idx, cell in enumerate(ws_mixup_sep[1], start=1) if cell.value == "NumSerial")
    actionid_col = next(col_idx for col_idx, cell in enumerate(ws_mixup_sep[1], start=1) if cell.value == "ActionID")
    motionid_col = next(col_idx for col_idx, cell in enumerate(ws_mixup_sep[1], start=1) if cell.value == "MotionID")

    for row in ws_mixup_sep.iter_rows(min_row=2):
        if row[numserial_col - 1].value:  # 检查值是否为None
            if row[actionid_col - 1].value:  # 检查值是否为None
                numserial_to_actionid[row[numserial_col - 1].value].append(str(row[actionid_col - 1].value))  # 将整数转换为字符串
            if row[motionid_col - 1].value:  # 检查值是否为None
                numserial_to_motionid[row[numserial_col - 1].value].append(str(row[motionid_col - 1].value))  # 将整数转换为字符串

    for key in numserial_to_actionid.keys():
        numserial_to_actionid[key] = [value for value in numserial_to_actionid[key] if str(value).lower() != 'nan']

    for key in numserial_to_motionid.keys():
        numserial_to_motionid[key] = [value for value in numserial_to_motionid[key] if str(value).lower() != 'nan']

    # 使用"NumSerial"列的值去匹配"ws_mixup_sep"工作表中的"NumSerial"
    for row in ws_mixup_reverse.iter_rows(min_row=2):
        numserial = row[0].value
        if numserial in numserial_to_actionid and numserial_to_actionid[numserial]:
            # 如果找到匹配的行，将"ws_mixup_sep"中"ActionID"列的值用英文逗号连接起来
            ws_mixup_reverse.cell(row=row[0].row, column=ws_mixup_reverse.max_column - 1, value=",".join(numserial_to_actionid[numserial]))
        else:
            ws_mixup_reverse.cell(row=row[0].row, column=ws_mixup_reverse.max_column - 1, value="")
        if numserial in numserial_to_motionid and numserial_to_motionid[numserial]:
            # 如果找到匹配的行，将"ws_mixup_sep"中"MotionID"列的值用英文逗号连接起来
            ws_mixup_reverse.cell(row=row[0].row, column=ws_mixup_reverse.max_column, value=",".join(numserial_to_motionid[numserial]))
        else:
            ws_mixup_reverse.cell(row=row[0].row, column=ws_mixup_reverse.max_column, value="")
    # endregion

    # region 格式&排序
    # 工作表排序
    sheet_names = wb.sheetnames
    mixup_index = sheet_names.index('MixUpReverse')# 获取'MixUp'工作表的索引
    sheet_names.insert(0, sheet_names.pop(mixup_index))# 将'MixUp'工作表移到最前面
    wb._sheets.sort(key=lambda sheet: sheet_names.index(sheet.title))# 重新排序工作表
    mixup_index = sheet_names.index('MixUpSep')# 获取'MixUp'工作表的索引
    sheet_names.insert(0, sheet_names.pop(mixup_index))# 将'MixUp'工作表移到最前面
    wb._sheets.sort(key=lambda sheet: sheet_names.index(sheet.title))# 重新排序工作表
    mixup_index = sheet_names.index('MixUp')# 获取'MixUp'工作表的索引
    sheet_names.insert(0, sheet_names.pop(mixup_index))# 将'MixUp'工作表移到最前面
    wb._sheets.sort(key=lambda sheet: sheet_names.index(sheet.title))# 重新排序工作表

    ConfigSheet(ws_mixup)
    ConfigSheet(ws_mixup_sep)
    ConfigSheet(ws_act2mot)
    ConfigSheet(ws_mot2event)
    ConfigSheet(ws_event2wemid)
    ConfigSheet(ws_mixup_reverse)
    # endregion

    wb.save(r'.\Yours\GBFR#0AllMixUp_PlayerVoice.xlsx')# 保存修改后的Excel文件

    # region 20：创建一个新的工作表"YourRecord"
    ws_yourRecord = wb.create_sheet("YourRecord", 0)

    # 在"YourRecord"工作表中创建列
    columns = ["ActionName_DIY", "NumSerial", "Alternative Voice Name", "Comment", "Reuse", "ActionID", "MotionID", "PL_vo_Event_WithSerial", "wemID", "bnk_Belonging"]
    for i, column in enumerate(columns, start=1):
        ws_yourRecord.cell(row=1, column=i, value=column)

    # 将"MixUp"工作表移动到"YourRecord"的右边
    mixup = wb['MixUp']
    wb._sheets.insert(1, wb._sheets.pop(wb.sheetnames.index('MixUp')))

    # 初始化一个空字符串来存储结果
    NumTips_str = ""

    # 遍历ws_event2wemid的每一行
    for row in ws_event2wemid.iter_rows(min_row=2, values_only=True):
        bnk_belonging = row[col_idx_bnkBelonging_E2i-1]
        num_serial = row[col_idx_NumSerial_E2i-1]

        # 从num_serial中提取字母
        letters = "".join(re.findall("[a-zA-Z]", num_serial))

        # 如果bnk_belonging不在字符串中，就添加它
        if bnk_belonging not in NumTips_str:
            NumTips_str += f"{bnk_belonging}: {letters.lower()}\n"

    # 在"YourRecord"工作表的第二行预设值
    ws_yourRecord.cell(row=2, column=ws_yourRecord['A1'].column).value = "[Example: KICK BACK]\n(If the value in column G is unique and the value in column F exists, it will be generated automatically.\nOtherwise, delete cells\' formula and fill in manually.)"
    ws_yourRecord.cell(row=2, column=ws_yourRecord['A1'].column).font = Font(color="00FF0000")
    ws_yourRecord.cell(row=2, column=ws_yourRecord['B1'].column).value = f"[Example: a_4]\n{NumTips_str}"
    ws_yourRecord.cell(row=2, column=ws_yourRecord['B1'].column).font = Font(color="00FF0000")
    ws_yourRecord.cell(row=2, column=ws_yourRecord['C1'].column).value = '[Your .wav Audio Files\' Name]\n(Don\'t include \".wav\")'
    ws_yourRecord.cell(row=2, column=ws_yourRecord['C1'].column).font = Font(color="00FF0000")
    ws_yourRecord.cell(row=2, column=ws_yourRecord['D1'].column).value = '[Just You Like]'
    ws_yourRecord.cell(row=2, column=ws_yourRecord['D1'].column).font = Font(color="00FF0000")
    ws_yourRecord.cell(row=2, column=ws_yourRecord['E1'].column).value = '[Auto]'
    ws_yourRecord.cell(row=2, column=ws_yourRecord['E1'].column).font = Font(color="000000FF")
    ws_yourRecord.cell(row=2, column=ws_yourRecord['F1'].column).value = '[Auto]'
    ws_yourRecord.cell(row=2, column=ws_yourRecord['F1'].column).font = Font(color="000000FF")
    ws_yourRecord.cell(row=2, column=ws_yourRecord['G1'].column).value = '[Auto]'
    ws_yourRecord.cell(row=2, column=ws_yourRecord['G1'].column).font = Font(color="000000FF")
    ws_yourRecord.cell(row=2, column=ws_yourRecord['H1'].column).value = '[Auto]'
    ws_yourRecord.cell(row=2, column=ws_yourRecord['H1'].column).font = Font(color="000000FF")
    ws_yourRecord.cell(row=2, column=ws_yourRecord['I1'].column).value = '[Auto]'
    ws_yourRecord.cell(row=2, column=ws_yourRecord['I1'].column).font = Font(color="000000FF")
    ws_yourRecord.cell(row=2, column=ws_yourRecord['J1'].column).value = '[Auto]'
    ws_yourRecord.cell(row=2, column=ws_yourRecord['J1'].column).font = Font(color="000000FF")

    # 在"YourRecord"工作表的第三行预设值
    ws_yourRecord.cell(row=3, column=ws_yourRecord['A1'].column).value = r'=IF(E3="Reuse","~Multi-Name",IFERROR(@INDEX(MixUp!$A:$A,MATCH(F3,MixUp!$B:$B,0)),"~Clear.Enter yourself"))'
    ws_yourRecord.cell(row=3, column=ws_yourRecord['B1'].column).value = 'a_4'
    ws_yourRecord.cell(row=3, column=ws_yourRecord['C1'].column).value = 'hum1'
    ws_yourRecord.cell(row=3, column=ws_yourRecord['D1'].column).value = ''
    ws_yourRecord.cell(row=3, column=ws_yourRecord['E1'].column).value = '=IF(LEN(G3)-LEN(SUBSTITUTE(G3,",",""))>=1,"Reuse","")'
    ws_yourRecord.cell(row=3, column=ws_yourRecord['F1'].column).value = '=INDEX(MixUpReverse!$1:$1048576,MATCH(B3,INDEX(MixUpReverse!$1:$1048576,,MATCH("NumSerial",MixUpReverse!$1:$1,0)),0),MATCH("ActionID",MixUpReverse!$1:$1,0))'
    ws_yourRecord.cell(row=3, column=ws_yourRecord['G1'].column).value = '=INDEX(MixUpReverse!$1:$1048576,MATCH(B3,INDEX(MixUpReverse!$1:$1048576,,MATCH("NumSerial",MixUpReverse!$1:$1,0)),0),MATCH("MotionID",MixUpReverse!$1:$1,0))'
    ws_yourRecord.cell(row=3, column=ws_yourRecord['H1'].column).value = '=INDEX(MixUpSep!$1:$1048576,MATCH(B3,INDEX(MixUpSep!$1:$1048576,,MATCH("NumSerial",MixUpSep!$1:$1,0)),0),MATCH("PL_vo_Event_WithSerial",MixUpSep!$1:$1,0))'
    ws_yourRecord.cell(row=3, column=ws_yourRecord['I1'].column).value = '=INDEX(MixUpSep!$1:$1048576,MATCH(B3,INDEX(MixUpSep!$1:$1048576,,MATCH("NumSerial",MixUpSep!$1:$1,0)),0),MATCH("wemID",MixUpSep!$1:$1,0))'
    ws_yourRecord.cell(row=3, column=ws_yourRecord['J1'].column).value = '=INDEX(MixUpSep!$1:$1048576,MATCH(B3,INDEX(MixUpSep!$1:$1048576,,MATCH("NumSerial",MixUpSep!$1:$1,0)),0),MATCH("bnk_Belonging",MixUpSep!$1:$1,0))'

    # "YourRecord"工作表的第一行加粗
    for cell in ws_yourRecord[1]:
        cell.font = Font(bold=True)

    # "YourRecord"工作表居中
    for row in ws_yourRecord.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 设置列宽
    # 遍历标题行的每一列
    for i, cell in enumerate(ws_yourRecord[1], start=1):
        # 检查单元格中是否有换行符
        if "\n" in str(cell.value):
            # 如果有换行符，那么找出最长的那一行
            lengths = [len(line) for line in str(cell.value).split("\n")]
            max_length = max(lengths)
        else:
            # 如果没有换行符，那么直接计算长度
            max_length = len(str(cell.value))
        # 设置列宽
        match get_column_letter(i):
            case 'A':
                ws_yourRecord.column_dimensions[get_column_letter(i)].width = max_length*2
            case 'B':
                ws_yourRecord.column_dimensions[get_column_letter(i)].width = max_length * 4
            case 'C':
                ws_yourRecord.column_dimensions[get_column_letter(i)].width = max_length * 2
            case 'D':
                ws_yourRecord.column_dimensions[get_column_letter(i)].width = max_length * 3
            case 'E':
                ws_yourRecord.column_dimensions[get_column_letter(i)].width = max_length * 1.5
            case 'F':
                ws_yourRecord.column_dimensions[get_column_letter(i)].width = max_length * 1.5
            case 'G':
                ws_yourRecord.column_dimensions[get_column_letter(i)].width = max_length * 1.5
            case 'H':
                ws_yourRecord.column_dimensions[get_column_letter(i)].width = max_length * 1.8
            case 'I':
                ws_yourRecord.column_dimensions[get_column_letter(i)].width = max_length * 2.5
            case 'J':
                ws_yourRecord.column_dimensions[get_column_letter(i)].width = max_length * 1.5
        # ws_yourRecord.column_dimensions[A].width = max_length*1.6

    # 冻结标题行
    ws_yourRecord.freeze_panes = ws_yourRecord['A2']
    # endregion

    wb.save(r'.\Yours\GBFR#0AllMixUp_PlayerVoice.xlsx')

    old_file_name = r".\Yours\GBFR#0AllMixUp_PlayerVoice.xlsx"
    new_file_name = fr".\Yours\GBFR#0{CharID}{CharName}VoiceInfo_MixUp.xlsx"

    # Check if the new file already exists
    if os.path.exists(new_file_name):
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        result = messagebox.askyesno("File Already Exists", "The new file already exists. Do you want to overwrite it?")
        root.destroy()
        if not result:
            # If the user chooses not to overwrite the new file, delete the old file
            if os.path.exists(old_file_name):
                os.remove(old_file_name)
            print("Operation cancelled.")
            return
        else:
            # If the user chooses to overwrite the new file, delete it
            os.remove(new_file_name)

    # Rename the file
    os.rename(old_file_name, new_file_name)
    print(f'Excel updated data saved in {new_file_name}')

if __name__ == "__main__":
    MixUp_Process()